# -*- coding: utf-8 -*-
import os
import time
import threading
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok
import cloudinary
from cloudinary.uploader import upload as cloudinary_upload, destroy as cloudinary_destroy
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import (
    MessageEvent, TextMessage, ImageMessage, TextSendMessage, ImageSendMessage, QuickReply, QuickReplyButton, MessageAction
)
from datetime import datetime, timedelta


# ========== 設定區 ==========
EXCEL_FILE = os.path.join("..", "SmartHome-main", "log.xlsx")
if not os.path.exists(EXCEL_FILE):
    print("❌ 找不到 log.xlsx")
    exit()
USER_ID_FILE = "user_ids.txt"
STATE_FILE = "last_row_num.txt"
USER_SETTINGS_FILE = "user_settings.txt"  # user_id,bind_name,img_url,public_id,mode

CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="
CHANNEL_SECRET = "7ed6db94815d685c613218f09978d655"  # 請替換為你的 LINE Channel Secret
CLOUDINARY_CLOUD_NAME = "dijfxhzqd"
CLOUDINARY_API_KEY = "158261511847545"
CLOUDINARY_API_SECRET = "cKny5zb9l2KhpoxIgUfKDOaTxlo"

# 初始化 Cloudinary 設定
cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True
)

# 初始化 LINE SDK
line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(CHANNEL_SECRET)

app = Flask(__name__)
last_row_num = 1  # 預設從第 2 行開始

# --------------------------------
# 工具函式區


# 🔍 自動尋找 log.xlsx 工具函式
def find_log_file(target_filename="log.xlsx", target_folder="SmartHome-main"):
    parent_dir = os.path.abspath(os.path.join(os.getcwd(), ".."))
    target_path = os.path.join(parent_dir, target_folder, target_filename)

    if os.path.exists(target_path):
        print(f"✅ 找到 log 檔案：{target_path}")
        return target_path
    else:
        print("❌ 無法找到 log.xlsx，請確認資料夾結構正確。")
        return None
def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w", encoding="utf-8") as f:
            pass
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            print("✅ 儲存新 userId：", user_id)

def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]
    

# 儲存用戶設定（包含綁定名字、圖片URL、public_id、模式）
def save_user_settings(user_id, bind_name, img_url, public_id, mode="圖文"):
    data = {}
    if os.path.exists(USER_SETTINGS_FILE):
        with open(USER_SETTINGS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                uid, name, url, pid, m = line.split(",", 4)
                data[uid] = (name, url, pid, m)

    # 如果已有舊圖片，要先刪除舊圖片 (如果 public_id 不同)
    if user_id in data:
        old_public_id = data[user_id][2]
        if old_public_id != public_id:
            try:
                cloudinary_destroy(old_public_id)
                print(f"✅ 已刪除舊圖片 public_id={old_public_id}")
            except Exception as e:
                print(f"❌ 刪除舊圖片失敗: {e}")

    data[user_id] = (bind_name, img_url, public_id, mode)

    with open(USER_SETTINGS_FILE, "w", encoding="utf-8") as f:
        for uid, (name, url, pid, m) in data.items():
            f.write(f"{uid},{name},{url},{pid},{m}\n")
    print(f"✅ 儲存 user({user_id}) 設定")

def load_user_settings():
    data = {}
    if os.path.exists(USER_SETTINGS_FILE):
        with open(USER_SETTINGS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                uid, name, url, pid, m = line.split(",", 4)
                data[uid] = {
                    "name": name,
                    "img_url": url,
                    "public_id": pid,
                    "mode": m
                }
    return data

def get_user_id_by_name(target_name):
    settings = load_user_settings()
    for uid, info in settings.items():
        if info["name"] == target_name:
            return uid
    return None

def push_message_text(text, user_id):
    try:
        line_bot_api.push_message(user_id, TextSendMessage(text=text))
        print(f"🔔 推播給 {user_id}：成功")
    except Exception as e:
        print(f"🔔 推播給 {user_id} 失敗：{e}")

def push_message_image(text, image_url, user_id):
    try:
        line_bot_api.push_message(user_id, [
            TextSendMessage(text=text),
            ImageSendMessage(original_content_url=image_url, preview_image_url=image_url)
        ])
        print(f"🔔 推播圖片給 {user_id}：成功")
    except Exception as e:
        print(f"🔔 推播圖片給 {user_id} 失敗：{e}")

def push_to_all_users(text, img_url=None):
    user_ids = load_user_ids()
    settings = load_user_settings()
    for uid in user_ids:
        user_mode = settings.get(uid, {}).get("mode", "圖文")
        if img_url and user_mode == "圖文":
            push_message_image(text, img_url, uid)
        else:
            push_message_text(text, uid)

def reply_quick_menu(reply_token):
    message = TextSendMessage(
        text="📋 請選擇操作項目：",
        quick_reply=QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="查詢最近紀錄", text="查詢最近紀錄")),
            QuickReplyButton(action=MessageAction(label="設定我的圖片與名字", text="我要設定圖片與名字")),
            QuickReplyButton(action=MessageAction(label="切換純文字/圖片模式", text="切換模式"))
        ])
    )
    line_bot_api.reply_message(reply_token, message)

def get_latest_entries_for_all():
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        last_records = {}

        # 從下往上讀，找每個人最後一筆非 Unknown 紀錄
        for row in range(max_row, 1, -1):
            values = [cell.value for cell in sheet[row][:3]]  # 讀前三欄: 時間、名字、狀態
            if len(values) < 2:
                continue
            time_raw, name = values[0], values[1]

            if not time_raw or not name:
                continue

            if name == "Unknown":
                continue  # 跳過陌生人

            # 時間格式化，假如是 datetime
            try:
                time_fmt = time_raw.strftime("%Y/%m/%d (%H:%M:%S)")
            except Exception:
                time_fmt = str(time_raw)

            if name not in last_records:
                last_records[name] = time_fmt

        if not last_records:
            return "⚠️ 找不到有效資料。"

        msgs = []
        for name, datetime_str in last_records.items():
            msgs.append(f"👤 {name}：\n📋 {datetime_str}")

        result_msg = "【Smart Home】\n\n為您列出各成員最後的回家紀錄：\n\n" + "\n\n".join(msgs)
        return result_msg

    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"



#________________陌生人函數
from datetime import datetime
from dateutil.parser import parse

def get_recent_unknown_entries(limit=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        unknown_entries = []
        now = datetime.now()
        for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
            time_str, name, result = row
            if name != "Unknown":
                continue
            try:
                log_time = parse(time_str)
            except Exception:
                continue
            if (now - log_time).days <= 7:
                unknown_entries.append((log_time, result))
            if len(unknown_entries) >= limit:
                break

        if not unknown_entries:
            return "【Smart Home】\n\n📭 最近 7 天內沒有陌生訪客記錄。"

        # 按時間由舊到新排序
        unknown_entries.sort(key=lambda x: x[0])

        msg = "【Smart Home】\n以下為陌生訪客的到訪記錄：\n\n"
        for log_time, result in unknown_entries:
            msg += f"🕒 {log_time.strftime('%Y/%m/%d %H:%M:%S')}\n📸 {result}\n\n"
        return msg.strip()
    except Exception as e:
        return f"【Smart Home】\n\n❌ 查詢時發生錯誤：{e}"






# --------------------------------
# Webhook 主流程

@app.route("/webhook", methods=["POST"])
def webhook():
    signature = request.headers.get("X-Line-Signature")
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        print("❌ Invalid signature. Please check the channel secret.")
        return "Invalid signature", 400
    return "OK"

@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    user_id = event.source.user_id
    reply_token = event.reply_token
    save_user_id(user_id)
    text = event.message.text.strip()

    if text == "選單":
        reply_quick_menu(reply_token)
    elif text == "查詢最近紀錄":
        msg = get_latest_entries_for_all()
        line_bot_api.reply_message(reply_token, TextSendMessage(text=msg))
    elif text == "查詢陌生人":
        msg = get_recent_unknown_entries()
        line_bot_api.reply_message(reply_token, TextSendMessage(text=msg))
    elif text == "我要設定圖片與名字":
        line_bot_api.reply_message(
            reply_token,
            TextSendMessage(
                text="【Smart Home】\n\n將為成員重設個人化圖片，請指定重設對象：\n\n提供名字"
            )
        )
        user_states[user_id] = {"step": "等待名字"}

    elif text == "切換模式":
        settings = load_user_settings()
        user_setting = settings.get(user_id)
        if not user_setting:
            line_bot_api.reply_message(
                reply_token,
                TextSendMessage(text="【Smart Home】\n\n尚未設定圖片與名字，請先設定。")
            )
        else:
            current_mode = user_setting["mode"]
            if current_mode == "圖文":
                new_mode = "純文字"
                msg = "【Smart Home】\n\n將屏蔽個人化圖片。"
            else:
                new_mode = "圖文"
                msg = "【Smart Home】\n\n將顯示個人化圖片。"

            save_user_settings(
                user_id,
                user_setting["name"],
                user_setting["img_url"],
                user_setting["public_id"],
                new_mode
            )
            line_bot_api.reply_message(
                reply_token,
                TextSendMessage(text=msg)
            )

    else:
        if user_id in user_states:
            step = user_states[user_id].get("step")
            if step == "等待名字":
                bind_name = text
                user_states[user_id]["bind_name"] = bind_name
                user_states[user_id]["step"] = "等待圖片"
                line_bot_api.reply_message(
                    reply_token,
                    TextSendMessage(
                        text=f"【Smart Home】\n\n已設定綁定名稱為：{bind_name}\n請接著傳送您要設定的圖片。"
                    )
                )
            else:
                line_bot_api.reply_message(
                    reply_token,
                    TextSendMessage(
                        text=f"【Smart Home】\n\n收到訊息：「{text}」\n請輸入「選單」查看操作項目。"
                    )
                )
        else:
            line_bot_api.reply_message(
                reply_token,
                TextSendMessage(
                    text=f"【Smart Home】\n\n收到訊息：「{text}」\n請輸入「選單」查看操作項目。"
                )
            )


@handler.add(MessageEvent, message=ImageMessage)
def handle_image_message(event):
    user_id = event.source.user_id
    reply_token = event.reply_token
    if user_id in user_states and user_states[user_id].get("step") == "等待圖片":
        message_id = event.message.id
        try:
            message_content = line_bot_api.get_message_content(message_id)
            image_data = message_content.content
            upload_result = cloudinary_upload(image_data)
            img_url = upload_result["secure_url"]
            public_id = upload_result["public_id"]

            bind_name = user_states[user_id].get("bind_name", "")
            save_user_settings(user_id, bind_name, img_url, public_id, "圖文")

            line_bot_api.reply_message(reply_token, TextSendMessage(
                text=f"【Smart Home】\n\n👤 {bind_name} 的個人化圖片已重新設定。"
            ))
            user_states.pop(user_id, None)
        except Exception as e:
            line_bot_api.reply_message(reply_token, TextSendMessage(
                text=f"❌ 圖片上傳失敗：{e}"
            ))
    else:
        line_bot_api.reply_message(reply_token, TextSendMessage(
            text="📤 請先使用「我要設定圖片與名字」指令開始設定。"
        ))

# --------------------------------
# 監控 Excel 新增資料並推播

def monitor_excel(interval=10):
    global last_row_num
    last_row_num = 1
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            try:
                last_row_num = int(f.read())
            except:
                last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ 找不到 {EXCEL_FILE}")
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_last_row = sheet.max_row

            if current_last_row > last_row_num:
                print(f"🆕 偵測到新增資料：{current_last_row - last_row_num} 列")
                for row in range(last_row_num + 1, current_last_row + 1):
                    values = [cell.value for cell in sheet[row][:3]]  # 讀前三欄: Time, Name, Result(可忽略)
                    if len(values) < 2 or any(v is None for v in values[:2]):
                        continue

                    time_str, name = values[0], values[1]

                    msg = f"【Smart Home】\n\n📋 新出入紀錄\n🕒 {time_str}\n👤 {name}"

                    # 用名字找對應的 user_id（用於判斷是否要附圖）
                    user_id_for_img = get_user_id_by_name(name)
                    settings = load_user_settings()
                    if user_id_for_img and user_id_for_img in settings:
                        img_url = settings[user_id_for_img].get("img_url")
                        if img_url:
                            push_to_all_users(msg, img_url)
                        else:
                            push_to_all_users(msg)
                    else:
                        push_to_all_users(msg)

                last_row_num = current_last_row
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    f.write(str(last_row_num))
        except Exception as e:
            print("❌ 監控錯誤：", e)
        time.sleep(interval)



# --------------------------------
# 簡易用戶狀態記錄 (用於設定流程)

user_states = {}

# --------------------------------
# 主程式入口

if __name__ == "__main__":
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}")
    print(f"📡 請設定 LINE Webhook 為：{public_url}/webhook")

    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)
